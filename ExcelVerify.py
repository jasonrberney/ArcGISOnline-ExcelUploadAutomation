import openpyxl
import os
import sys
print os.getcwd()
import warnings
warnings.filterwarnings("ignore")

def analyzeWB(wb_input):
    try:
        # open the Excel workbook and get the active sheet
        wb = openpyxl.load_workbook(wb_input)
        mainSheet = wb.get_sheet_names()
        ws = wb.get_active_sheet()
        print ws

        # get number of columns and rows in Excel file
        columnQuant = ws.get_highest_column()
        print "There are " + str(columnQuant) + " columns in your excel file."
        rowQuant = ws.get_highest_row() - 1
        print "There are " + str(rowQuant) + " rows of data in your excel file."

        # create list of all possible names for latitude and longitude
        latitudeNames = ["lat", "phi", "latitude", "y", "ycoordinate", "ycoord", "y_coordinate"]
        longitudeNames = ["lng", "lon", "lambda", "long", "longitude", "x", "xcoordinate", "xcoord", "x_coordinate"]
        # find values of first row in workbook
        for cellObj in ws.rows[0]:
            cellObjValue = cellObj.value
            # confirm that there are columns for latitude and longitude data
            if str(cellObjValue).lower() in latitudeNames:
                latitudeLoc = str(cellObj)
                print "Excel has latitude data in " + latitudeLoc
            elif str(cellObjValue).lower() in longitudeNames:
                longitudeLoc = str(cellObj)
                print "Excel has longitude data in " + longitudeLoc
            elif cellObjValue == None:
                print "Uh Oh! One of your column headings is empty! Please edit the spreadsheet and rerun the script."
                sys.exit("Please edit your spreadsheet by adding a column header.")
                break
            else:
                pass

        # Check every latitude and longitude for a value and notify user if value is missing
        for row in range(2, ws.get_highest_row() + 1):
            latitude  = ws[latitudeLoc[11] + str(row)].value
            longitude = ws[longitudeLoc[11] + str(row)].value
            if latitude is None:
                latAnswer = raw_input("you have an empty latitude value in row " + str(row) + ", would you like to stop the process and make an edit? (yes or no)")
                if str(latAnswer).lower() == "yes":
                    sys.exit()
                else:
                    continue
            elif longitude is None:
                lngAnswer = raw_input("you have an empty longitude value in row " + str(row) + ", would you like to stop the process and make an edit? (yes or no)")
                if str(lngAnswer).lower() == "yes":
                    sys.exit()
                else:
                    continue

    except UserWarning:
        print "User warning, likely for calling to deprecated function or class..."


if __name__ == "__main__":
    analyzeWB("C:\Users\Jason\Documents\MSGT\GIS Automation & Customization\Final Project\MOCK_DATA_Error.xlsx")